import json
import re
import unicodedata
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\40775791\Downloads\CRONOGRAMA HMSA 05.12 RDC.xlsx")
DEST = ROOT / "prototipo" / "src" / "data" / "cronogramaHmsa.js"

MODEL_ID = "hmsa-2026"
FONTE = "CRONOGRAMA HMSA 05.12 RDC.xlsx"
HOSPITAL = "Hospital Municipal Souza Aguiar"
SIGLA = "HMSA"

ABAS_IGNORADAS = {
    "Histórico de Versões",
    "REDUÇÃO",
    "Base Salarial RH",
    "RESUMO",
    "CRONOGRAMA HMSA",
    "CRONOGRAMA HMSA CEBAS",
}


def aba_de_equipe(nome):
    return bool(re.match(r"^\s*b\d+\s*[-–]", str(nome or ""), flags=re.IGNORECASE))


def normalizar(valor):
    texto = str(valor or "")
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = re.sub(r"\s+", " ", texto).strip().lower()
    return texto


def slug(valor):
    texto = normalizar(valor)
    texto = re.sub(r"[^a-z0-9]+", "-", texto).strip("-")
    return texto or "item"


def limpar_texto(valor):
    return re.sub(r"\s+", " ", str(valor or "")).strip()


def numero(valor):
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return float(valor)
    if valor in (None, ""):
        return None
    texto = str(valor).strip()
    if not texto or texto == "-":
        return None
    texto = texto.replace("R$", "").replace(" ", "")
    try:
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        return float(texto)
    except ValueError:
        return None


def numero_ou_zero(valor):
    n = numero(valor)
    return n if n is not None else 0


def indice(celulas, teste):
    for idx, valor in enumerate(celulas):
        if teste(normalizar(valor)):
            return idx
    return -1


def mapear_cabecalho(linhas):
    melhor = {"linha": -1, "pontos": -1, "valores": []}
    for idx, linha in enumerate(linhas[:25]):
        textos = [normalizar(celula) for celula in linha]
        pontos = sum(
            [
                any("categoria" in t for t in textos),
                any("carga horaria" in t for t in textos),
                any("quantitativo" in t for t in textos),
                any("salario" in t or "remuneracao" in t for t in textos),
            ]
        )
        if pontos > melhor["pontos"]:
            melhor = {"linha": idx, "pontos": pontos, "valores": linha}

    h = melhor["valores"]
    remuneracao_bruta = indice(h, lambda t: "remuneracao bruta" in t)
    totais = [
        {"idx": i, "texto": normalizar(valor)}
        for i, valor in enumerate(h)
        if "salario total" in normalizar(valor)
    ]
    salario_total = next(
        (item["idx"] for item in totais if item["idx"] > remuneracao_bruta),
        totais[0]["idx"] if totais else -1,
    )

    return {
        "linha": melhor["linha"],
        "categoria": indice(h, lambda t: "categoria" in t),
        "chs": indice(h, lambda t: "carga horaria" in t),
        "quantidade": indice(h, lambda t: "quantitativo" in t and "turno" not in t),
        "quantidadeTurno": indice(h, lambda t: "quantitativo" in t and "turno" in t),
        "base": indice(h, lambda t: "salario" in t and "base" in t),
        "insalubridade": indice(h, lambda t: "insalubridade" in t),
        "gratificacao": indice(h, lambda t: "gratificacao" in t),
        "titulacao": indice(h, lambda t: "titulacao" in t),
        "adicionalNoturno": indice(h, lambda t: "adic" in t and "noturno" in t),
        "remuneracaoBruta": remuneracao_bruta,
        "salarioTotal": salario_total,
        "valeRefeicao": indice(h, lambda t: "vale refeicao" in t or "vale alimentacao" in t),
        "valeTransporte": indice(h, lambda t: "vale transporte" in t),
        "diasOperacionais": indice(h, lambda t: "dias operacionais" in t),
    }


def valor_linha(linha, col):
    if col < 0 or col >= len(linha):
        return None
    return linha[col]


def n_linha(linha, col):
    return numero(valor_linha(linha, col))


def n0_linha(linha, col):
    return numero_ou_zero(valor_linha(linha, col))


def extrair_setor(ws):
    linhas = list(ws.iter_rows(values_only=True))
    colunas = mapear_cabecalho(linhas)
    cat_col = colunas["categoria"]
    if colunas["linha"] < 0 or cat_col < 0:
        return None

    setor_id = slug(ws.title)
    grupos = []
    linhas_equipe = []
    usados_grupo = set()

    for row_idx in range(colunas["linha"] + 1, len(linhas)):
        linha = linhas[row_idx]
        nome = limpar_texto(valor_linha(linha, cat_col))
        if not nome:
            continue
        norm = normalizar(nome)
        if norm.startswith("total"):
            break
        if norm.startswith("encargos sociais") or norm == "beneficios":
            break

        chs = n_linha(linha, colunas["chs"])
        qtd = n_linha(linha, colunas["quantidade"])
        if chs is not None and qtd is not None:
            linha_id = f"{setor_id}-l{len(linhas_equipe) + 1}"
            linhas_equipe.append(
                {
                    "id": linha_id,
                    "categoria": nome,
                    "perfilId": f"modelo:{slug(nome)}",
                    "chs": int(chs) if chs.is_integer() else chs,
                    "quantidade": int(qtd) if qtd.is_integer() else qtd,
                    "quantidadeTurno": n0_linha(linha, colunas["quantidadeTurno"]),
                    "revisar": False,
                    "linhaOrigem": row_idx + 1,
                }
            )
            continue

        ignorar_grupo = any(
            trecho in norm
            for trecho in [
                "custeio de pessoal",
                "aplicacao de recursos",
                "inss",
                "fgts",
                "pis",
                "sub total",
                "total provisoes",
                "total encargos",
                "vale transporte",
                "vale refeicao",
                "aplicacao dos recursos",
            ]
        )
        if ignorar_grupo:
            continue

        grupo_id_base = slug(nome)
        grupo_id = grupo_id_base
        contador = 2
        while grupo_id in usados_grupo:
            grupo_id = f"{grupo_id_base}-{contador}"
            contador += 1
        usados_grupo.add(grupo_id)
        grupos.append({"id": grupo_id, "nome": nome.upper(), "linhaOrigem": row_idx + 1})

    componentes = {}
    for item in linhas_equipe:
        linha = linhas[item["linhaOrigem"] - 1]
        componentes[item["id"]] = {
            "chs": item["chs"],
            "quantidade": item["quantidade"],
            "quantidadeTurno": n0_linha(linha, colunas["quantidadeTurno"]),
            "base": n0_linha(linha, colunas["base"]),
            "insalubridade": n0_linha(linha, colunas["insalubridade"]),
            "gratificacao": n0_linha(linha, colunas["gratificacao"]),
            "titulacao": n0_linha(linha, colunas["titulacao"]),
            "adicionalNoturno": n0_linha(linha, colunas["adicionalNoturno"]),
            "remuneracaoBruta": n0_linha(linha, colunas["remuneracaoBruta"]),
            "salarioTotal": n0_linha(linha, colunas["salarioTotal"]),
            "valeRefeicaoBeneficiarios": n0_linha(linha, colunas["valeRefeicao"]),
            "valeTransporteBeneficiarios": n0_linha(linha, colunas["valeTransporte"]),
            "diasOperacionais": n0_linha(linha, colunas["diasOperacionais"]),
        }

    return {
        "setor": {
            "id": setor_id,
            "nome": ws.title,
            "abaOrigem": ws.title,
            "linhas": linhas_equipe,
            "resumo": {
                "linhasEquipe": len(linhas_equipe),
                "profissionais": sum(float(item["quantidade"]) for item in linhas_equipe),
                "revisar": sum(1 for item in linhas_equipe if item["revisar"]),
            },
        },
        "agrupadores": grupos,
        "componentes": componentes,
    }


def extrair_resumo_financeiro(wb):
    ws = wb["RESUMO"]
    sem_ano1 = numero_ou_zero(ws["O7"].value)
    sem_ano2 = numero_ou_zero(ws["O10"].value)
    sem_contrato = numero_ou_zero(ws["P10"].value)
    com_ano1 = numero_ou_zero(ws["O24"].value)
    com_ano2 = numero_ou_zero(ws["O27"].value)
    com_contrato = numero_ou_zero(ws["P27"].value)
    reducao = sem_contrato - com_contrato
    return {
        "status": "extraido",
        "fonteAba": "RESUMO",
        "periodoMeses": 24,
        "semCebas": {
            "mes1": numero_ou_zero(ws["C7"].value),
            "ano1": sem_ano1,
            "ano2": sem_ano2,
            "contrato": sem_contrato,
        },
        "comCebas": {
            "mes1": numero_ou_zero(ws["C24"].value),
            "ano1": com_ano1,
            "ano2": com_ano2,
            "contrato": com_contrato,
        },
        "reducao": {
            "valor": reducao,
            "percentual": reducao / sem_contrato if sem_contrato else 0,
        },
        "observacao": "Valores históricos extraídos da planilha de origem. Ao criar um plano, o aplicativo recalcula o financeiro com a base vigente.",
    }


def extrair_cronograma_historico(ws):
    grupos_ordem = []
    grupos = {}
    linhas = {}
    ano = 1
    grupo_atual = None

    def obter_grupo(titulo):
        gid = slug(titulo)
        if gid not in grupos:
            grupos[gid] = {"id": gid, "titulo": titulo, "linhas": []}
            grupos_ordem.append(gid)
        return gid

    for row_idx in range(1, ws.max_row + 1):
        label = limpar_texto(ws.cell(row_idx, 1).value)
        norm = normalizar(label)
        linha_vals = [ws.cell(row_idx, col).value for col in range(1, 18)]
        textos = [normalizar(v) for v in linha_vals if isinstance(v, str)]
        if any("mes 13" in t for t in textos):
            ano = 2
        elif any("mes 01" in t for t in textos):
            ano = 1

        if norm.startswith("parte 1"):
            grupo_atual = obter_grupo("PARTE 1 - APOIO À GESTÃO")
            continue
        if norm.startswith("parte 2"):
            grupo_atual = obter_grupo("PARTE 2 - RH E CUSTEIO")
            continue
        if norm.startswith("parte 3"):
            grupo_atual = obter_grupo("PARTE 3 - INVESTIMENTO - ADAPTAÇÕES E EQUIPAMENTOS")
            continue
        if norm.startswith("parte 4"):
            grupo_atual = obter_grupo("PARTE 4 - VARIÁVEL")
            continue
        if not label or norm in {"item", "cronograma de desembolso hmsa", "cronograma de desembolso - hmsa"}:
            continue
        if norm.startswith("cronograma de desembolso") or norm.startswith("parcela "):
            continue

        valores = [numero_ou_zero(ws.cell(row_idx, col).value) for col in range(4, 16)]
        if not any(valores):
            continue

        grupo_linha = obter_grupo("CRONOGRAMA DE DESEMBOLSO - HMSA") if norm.startswith("total parte") or norm.startswith("total termo") else grupo_atual
        if not grupo_linha:
            continue

        key = f"{grupo_linha}:{slug(label)}"
        if key not in linhas:
            tipo = "item"
            if norm.startswith("total termo"):
                tipo = "grand-total"
            elif norm.startswith("total") or norm.startswith("d- total") or norm.startswith("parcela fixa"):
                tipo = "subtotal"
            elif norm.startswith("a - apoio") or norm.startswith("b 1 - rh") or norm.startswith("b - rh") or norm.startswith("d - investimento") or norm.startswith("variavel"):
                tipo = "total"
            linhas[key] = {
                "id": slug(label),
                "label": label,
                "tipo": tipo,
                "valorUnitario": numero_ou_zero(ws.cell(row_idx, 3).value),
                "valores": [0] * 24,
                "fonte": ws.title,
                "memoria": "Linha histórica extraída da aba final da planilha de origem.",
            }
            grupos[grupo_linha]["linhas"].append(linhas[key])

        inicio = 0 if ano == 1 else 12
        for i, valor in enumerate(valores):
            linhas[key]["valores"][inicio + i] = valor
        if ano == 1:
            total_ano1 = numero(ws.cell(row_idx, 16).value)
            if total_ano1 is not None:
                linhas[key]["totalAno1"] = total_ano1
        else:
            total_ano2 = numero(ws.cell(row_idx, 16).value)
            total_contrato = numero(ws.cell(row_idx, 17).value)
            if total_ano2 is not None:
                linhas[key]["totalAno2"] = total_ano2
            if total_contrato is not None:
                linhas[key]["totalContrato"] = total_contrato

    for linha in linhas.values():
        if not linha["valorUnitario"]:
            linha["valorUnitario"] = next((v for v in linha["valores"] if v), 0)
        linha["totalAno1"] = linha.get("totalAno1", sum(linha["valores"][:12]))
        linha["totalAno2"] = linha.get("totalAno2", sum(linha["valores"][12:24]))
        linha["totalContrato"] = linha.get("totalContrato", linha["totalAno1"] + linha["totalAno2"])

    return [grupos[gid] for gid in grupos_ordem if grupos[gid]["linhas"]]


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    setores = []
    agrupadores = {}
    componentes = {}
    perfis = {}

    for ws in wb.worksheets:
        if ws.title in ABAS_IGNORADAS or not aba_de_equipe(ws.title):
            continue
        extraido = extrair_setor(ws)
        if not extraido or not extraido["setor"]["linhas"]:
            continue
        setor = extraido["setor"]
        setores.append(setor)
        agrupadores[setor["id"]] = extraido["agrupadores"]
        componentes[setor["id"]] = extraido["componentes"]
        for linha in setor["linhas"]:
            perfis[linha["perfilId"]] = {
                "id": linha["perfilId"],
                "nome": linha["categoria"],
                "categoria": linha["categoria"],
                "funcao": linha["categoria"],
                "revisar": False,
            }

    linhas = [linha for setor in setores for linha in setor["linhas"]]
    resumo_financeiro = extrair_resumo_financeiro(wb)
    resumo_financeiro["semCebas"]["grupos"] = extrair_cronograma_historico(wb["CRONOGRAMA HMSA"])
    resumo_financeiro["comCebas"]["grupos"] = extrair_cronograma_historico(wb["CRONOGRAMA HMSA CEBAS"])
    modelo = {
        "id": MODEL_ID,
        "nome": "Cronograma Hospital Municipal Souza Aguiar",
        "hospitalNome": HOSPITAL,
        "unidadeModelo": HOSPITAL,
        "sigla": SIGLA,
        "grupoId": "SUE",
        "grupoNome": "Hospitais de Urgência e Emergência",
        "descricao": "Modelo clonável baseado no cronograma completo do Hospital Municipal Souza Aguiar, com serviços, categorias gerais e equipes importadas como se tivessem sido montadas no aplicativo.",
        "fonte": FONTE,
        "tipo": "modelo_cronograma",
        "status": "validado",
        "modoUso": "clonavel",
        "observacao": "Modelo clonável: estrutura e equipe importadas da planilha; componentes salariais preservados e financeiro recalculado pelo app.",
        "parametrosCronograma": {
            "custeioOperacionalPct": numero_ou_zero(wb["RESUMO"]["J2"].value),
            "valeTransporteDia": numero_ou_zero(wb["RESUMO"]["G2"].value),
            "valeRefeicaoDia": numero_ou_zero(wb["RESUMO"]["G3"].value),
            "salarioMinimo": numero_ou_zero(wb["RESUMO"]["D2"].value),
            "mesesCronograma": 24,
        },
        "abasIgnoradas": sorted(ABAS_IGNORADAS),
        "abasOrigem": [setor["abaOrigem"] for setor in setores],
        "setores": setores,
        "resumoFinanceiro": resumo_financeiro,
        "resumo": {
            "setores": len(setores),
            "linhasEquipe": len(linhas),
            "profissionais": sum(float(linha["quantidade"]) for linha in linhas),
            "linhasRevisao": sum(1 for linha in linhas if linha["revisar"]),
        },
    }

    payload = (
        "// Gerado por scripts/gerar-cronograma-hmsa.py.\n"
        "// Modelo clonável do cronograma completo do Hospital Municipal Souza Aguiar.\n\n"
        f"export const perfisCronogramaHmsa = {json.dumps(list(perfis.values()), ensure_ascii=False, indent=2)}\n\n"
        f"export const cronogramasProntosHmsa = {json.dumps([modelo], ensure_ascii=False, indent=2)}\n\n"
        f"export const agrupadoresCronogramaHmsa = {json.dumps({MODEL_ID: agrupadores}, ensure_ascii=False, indent=2)}\n\n"
        f"export const componentesPlanilhaCronogramaHmsa = {json.dumps({MODEL_ID: componentes}, ensure_ascii=False, indent=2)}\n"
    )
    DEST.write_text(payload, encoding="utf-8")
    print(
        json.dumps(
            {
                "destino": str(DEST),
                "setores": len(setores),
                "linhasEquipe": len(linhas),
                "profissionais": modelo["resumo"]["profissionais"],
                "perfis": len(perfis),
                "custeioPct": modelo["parametrosCronograma"]["custeioOperacionalPct"],
                "semCebasContrato": resumo_financeiro["semCebas"]["contrato"],
                "comCebasContrato": resumo_financeiro["comCebas"]["contrato"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
